# -*- coding: utf-8 -*-
# -*- ver	: >=3.8.0 -*-
# -*- coder : CO0kie丶 -*-
# -*- time  : 20201119 -*-
import pprint
from openpyxl import Workbook

pp = pprint.PrettyPrinter()


def AddSheet(wb, SheetName: str, Dict: dict = None, List: list = None):
    ws = wb.create_sheet(SheetName)
    if List is None:
        for key in Dict:
            ws.append((key, Dict[key]))
    else:
        for rows in List:
            ws.append(rows)
    ws.close()
    pass


def SaveXlsx2(Path: str, Dict):
    for url in Dict:
        wb = Workbook(write_only=True)

        less = url.find('Less')
        less = url[less:url.find('/', less)]
        print(f'\n【{url}】{less}')
        # pp.pprint(Dict[url])

        for key in Dict[url]:
            print(key)
            pp.pprint(Dict[url][key])

            if "注入方式" == key:
                AddSheet(wb, key, Dict[url][key])
            elif "脱库" == key:
                for key2 in Dict[url][key]:
                    if '数据库' == key2:
                        AddSheet(wb, key2, Dict[url][key][key2])
                    elif '数据表' == key2:
                        for tb_name in Dict[url][key][key2]:
                            if 'data' in Dict[url][key][key2][tb_name]:
                                AddSheet(wb, tb_name, List=Dict[url][key][key2][tb_name]['data'])
                            else:
                                AddSheet(wb, tb_name, List=Dict[url][key][key2][tb_name])
                        pass
                    pass
                pass
            pass
        wb.save(f'{Path}{less}.xlsx')
    pass


# noinspection PyBroadException
def SaveXlsx(FileName: str, Dict):
    """
    函数： 保存Dict里的值

    :param FileName: 文件路径
    :param Dict: 传入的漏洞信息
    :return: T/F=是否保存成功
    """

    # 保存xlsx
    wb = Workbook(write_only=True)

    ws1 = wb.create_sheet('数据库信息')

    for data in Dict:
        if '注入方式' == data:
            continue

        # 排版信息
        if '脱裤_数据库信息' == data:
            for data2 in Dict[data]:
                ws1.append((data2, Dict[data][data2]))
        elif '脱裤_数据库数据' == data:
            for tb_name in Dict[data]:
                ws = wb.create_sheet(tb_name)
                for line in Dict[data][tb_name]['data']:
                    ws.append(line)
                ws.close()
        elif '注入' in data:
            ws1.append((data, Dict[data]))
            pass

    # print(Dict)
    print(FileName)
    wb.save(FileName)
    wb.close()
    pass
